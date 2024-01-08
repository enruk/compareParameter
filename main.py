import pandas as panda 
import tkinter as tk
import os
import openpyxl 
import gui
from parameter import parameter_list
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
     
class excel_file:
    def __init__(self,name,project_list):
        self.name = name
        self.file_path = ""
        self.folder_path = ""
        self.project_list = project_list 
        
    def set_target_file_path(self): 
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        
        self.folder_path = os.path.join(desktop_path, "Parameter Comparison")
        if not os.path.exists(self.folder_path):
            os.mkdir(self.folder_path)
        self.file_path = os.path.join(self.folder_path, f"{self.name}.xlsx")              

    def write_first_project_to_excel(self):
        
        row_a = []
        row_b = []
        
        for parameter in self.project_list[0].standard_template.parameters:
            row_a.append(parameter.name)
            row_b.append(parameter.value)
        
        df = panda.DataFrame({'Parameter': row_a, self.project_list[0].name: row_b})
        df.to_excel(self.file_path,index=False)  
        
    def write_local_changes_to_excel(self,index_in_project_list):
        
        row_a = []
        row_b = []
        
        for parameter in self.project_list[index_in_project_list].local_parameter_changes.parameters:
            row_a.append(parameter.name)
            row_b.append(parameter.value)
        
        df = panda.DataFrame({'Parameter': row_a, self.project_list[index_in_project_list].name: row_b})
        df.to_excel(self.file_path,index=False)  
        
    def add_next_project(self,index_in_project_list):
        
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook.active
        sheet.cell(row=1, column=self.project_list[index_in_project_list].column_in_excel, value=self.project_list[index_in_project_list].name)
        
        for parameter in self.project_list[index_in_project_list].standard_template.parameters:
            look_up_parameter = parameter.name
            counter = 0
            look_up_parameter_found = False
            for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row, values_only=True):
                counter = counter + 1
                if row[0] == look_up_parameter:
                    sheet.cell(row=counter, column=self.project_list[index_in_project_list].column_in_excel, value=parameter.value)
                    look_up_parameter_found = True
                    break
            
            # if parameter wasnt found add it at the end
            if not look_up_parameter_found:
                next_free_row = sheet.max_row+1
                sheet.cell(row=next_free_row, column=1, value=parameter.name)
                sheet.cell(row=next_free_row, column=self.project_list[index_in_project_list].column_in_excel, value=parameter.value)
    
        workbook.save(self.file_path)
        
    def compare_parameters_to_master_parameters(self,index_in_project_list):
        
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook.active
        
        counter = 2
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            
            value_row_master = str(row[1]) 
            value_row_x = str(row[self.project_list[index_in_project_list].column_in_excel-1])  

            if value_row_master != value_row_x:
                    cell = sheet.cell(row=counter,column = self.project_list[index_in_project_list].column_in_excel)
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            counter = counter + 1
        
        workbook.save(self.file_path)
        
    def format_sheet(self):
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook.active
        sheet.insert_cols(2, amount=2)
        
        counter = 2
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            if row[0] and "." in row[0]:
                parts = row[0].split(".")
                sheet.cell(row=counter, column=1).value = parts[0] if len(parts) >= 1 else ""
                sheet.cell(row=counter, column=2).value = parts[1] if len(parts) >= 2 else ""
                sheet.cell(row=counter, column=3).value = parts[2] if len(parts) >= 3 else ""
            counter = counter + 1
            
        for column_letter in "ABCDEFGH":
            sheet.column_dimensions[column_letter].width = 50
            
        font = Font(bold=True)
        alignment = Alignment(horizontal='center')
        border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

        header = sheet[1]
        # Setze die Formatierung für jede Zelle in der oberen Zeile
        for cell in header:
            cell.font = font
            cell.alignment = alignment
            cell.border = border

        workbook.save(self.file_path)

                

class project:
    def __init__(self,name,folder_path_project,column_in_excel):
        self.name = name
        
        self.path_to_PLC_folder = folder_path_project
        self.path_to_project_special_folder = ""
        self.path_to_project_folder= ""
        self.file_path_standard_template = ""
        self.file_path_standard_templates_changes = ""
        
        self.standard_template = parameter_list
        self.standard_templates_changes = parameter_list
        self.local_parameter_changes = parameter_list
        self.column_in_excel = column_in_excel
        
    def get_path_of_templates(self):
        
        self.path_to_project_special_folder = self.path_to_PLC_folder + "/OverheadSystems/ProjectSpecials" 
        
        name_standardtemplate = "PRG_StandardTemplate.TcPOU"
        name_standardtemplateschanges = "PRG_StandardTemplatesChanges.TcPOU"
         
        files = os.listdir(self.path_to_project_special_folder)
                
        if name_standardtemplate in files and name_standardtemplateschanges in files:
            self.file_path_standard_template = self.path_to_project_special_folder + "/" +name_standardtemplate
            self.file_path_standard_templates_changes = self.path_to_project_special_folder + "/" +name_standardtemplateschanges
        else:
            print(f"For the projekt '{self.name}' it was not possible to find the PRG_StandardTemplate.TcPOU and / or the PRG_StandardTemplatesChanges.TcPOU")

    def get_local_changes(self):
        
        self.local_parameter_changes = parameter_list("dummy_path")
        
        # go to project folder and get all folders of the PLC programm
        self.path_to_project_folder = self.path_to_PLC_folder + "/OverheadSystems/Project"
        content_project_folder = os.listdir(self.path_to_project_folder)
        content_not_to_check = ["Hardware", "Area50", "General", "GVLs","MAIN.TcPOU"]
        clean_project_folder = [folder for folder in content_project_folder if folder not in content_not_to_check]
        
        # go in every sub folders and get all programs in each sub folder
        for sub_folder in clean_project_folder:
            temp_sub_folder_path = self.path_to_project_folder + "/" + sub_folder
            content_program_folder = os.listdir(temp_sub_folder_path)
            clean_program_folder = [element for element in content_program_folder if element != "DUTs"]
            
            # go throw all programs, create a temporaly list of local changes
            for program in clean_program_folder:
                program_path = temp_sub_folder_path + "/" + program
                temp_parameter_list = parameter_list(program_path)
                temp_parameter_list.read_local_param_from_file()
                
                # copy all temporary changes to the main list for local changes
                self.local_parameter_changes.parameters.extend(temp_parameter_list.parameters)
                temp_parameter_list.parameters.clear()
    
    def fill_parameter_lists_from_xml(self):
        self.standard_template = parameter_list(self.file_path_standard_template)
        self.standard_template.read_param_from_file()
        
        self.standard_templates_changes = parameter_list(self.file_path_standard_templates_changes)
        self.standard_templates_changes.read_param_from_file()
    
    def write_changes_to_standard_template(self):
        for parameter_changes in self.standard_templates_changes.parameters:
            counter = 0
            for parameter_standard in self.standard_template.parameters:
                if parameter_standard.name == parameter_changes.name:
                    self.standard_template.parameters[counter].value = parameter_changes.value
                    break
                counter = counter + 1       
                
                
                
if __name__ == "__main__":
    
    root = tk.Tk()
    app = gui.user_interface(root)
    root.mainloop()
    
    # format data_array from gui to project_data[X][Y], X project, Y data (0=name, 1=folder path)
    app.data_array = list(filter(lambda x: x != "", app.data_array))
    project_data = []
    project_data = [app.data_array[i:i+2] for i in range(0, len(app.data_array), 2)]  
    
    # read projects
    project_list = []
    number_of_projects = int(len(project_data))
    for index_in_project_list in range(number_of_projects):
        # create new project
        project_name = project_data[index_in_project_list][0]
        project_folder_path = project_data[index_in_project_list][1]
        column_in_excel = index_in_project_list + 2 #column 1=parameters, therefore first project[0] in column = 2
        new_project = project(project_name,project_folder_path,column_in_excel) 
        
        # get file pathes of standardTemplate and StandardTemplatesChanges
        new_project.get_path_of_templates()
        
        # fill parameter_list with data from templates
        new_project.fill_parameter_lists_from_xml()
        
        # ove#rwrite parameter from standardtemplateschanges to standardtemplate
        new_project.write_changes_to_standard_template()
        
        # get the local changes
        new_project.get_local_changes()
        
        # add project to projectlist
        project_list.append(new_project)
    
    # create excel file and get desktop path
    excel = excel_file("Parameter Comparison",project_list)
    excel.set_target_file_path()
    
    # write first project to excel
    excel.write_first_project_to_excel()
    
    # add more projects to excel sheet
    if number_of_projects > 1:
        for index_in_project_list in range(1,number_of_projects):
            excel.add_next_project(index_in_project_list)
            excel.compare_parameters_to_master_parameters(index_in_project_list)
        
    for index_in_project_list in range(number_of_projects):
        excel_local_changes = excel_file(project_list[index_in_project_list].name,project_list)
        excel_local_changes.set_target_file_path()
        excel_local_changes.write_local_changes_to_excel(index_in_project_list)
            

    # Add 2 columns between A and B and split GVL to make it easier to filter in excel
    excel.format_sheet()
    
    # Tell user script is done
    print("Compare Done")
    