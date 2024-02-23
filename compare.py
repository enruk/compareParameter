import pandas as panda 
import os
import openpyxl 
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import xml.etree.ElementTree as ET
from copy import copy
     
class comparison:
    
    def __init__(self):
        
        self.path_target_folder = ""
        self.path_file_comparison = ""
        self.path_file_filtered_comparion = ""
        self.path_file_differences = ""
        
        self.project_list = []
   
        self.ignored_parameters = []
        self.Ignored_folders = []
        
        self.summary_differnces_On = False
        self.local_changes_On = False
        self.filter = []
        self.filters_On = False
        self.filter_index = 0, # 1=induction, 2=dynamicbuffer, 3=orderbuffer, 4=matrix_presorter, 5=packing

             
    def get_file_pathes(self): 
        # get the file pathes to create and save the excel files
        
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        
        self.path_target_folder = os.path.join(desktop_path, "Parameter Comparison")
        if not os.path.exists(self.path_target_folder):
            os.mkdir(self.path_target_folder)
        self.path_file_comparison = os.path.join(self.path_target_folder, "Parameter Comparison.xlsx")     
        self.path_file_filtered_comparion = os.path.join(self.path_target_folder, "Parameter Comparison - Fitlered.xlsx")     
        self.path_file_differences = os.path.join(self.path_target_folder, "Parameter Comparison - Differences.xlsx")
    
    
    def get_new_working_path(self,original_path):
        # get the real working path of the script, this is needed to get access to the xml's
        
        directory, filename = os.path.split(original_path)
        folders = directory.split(os.path.sep)
        if 'dist' in folders:
            folders.remove('dist') 
        new_directory = os.path.sep.join(folders)
        new_path = f'{new_directory}{os.path.sep}{filename}'

        return new_path
        
        
    def get_info_for_script(self):
        # get info for ignored foldes and parameters from the XML's

        # get path to xml
        original_working_dir = os.getcwd()
        new_path = self.get_new_working_path(original_working_dir)
        xml_file_path = os.path.join(new_path, 'stuff_to_ignore.xml')

        # parse the XML data
        tree = ET.parse(xml_file_path)
        wurzel = tree.getroot()

        # find IgnoredFolders
        self.Ignored_folders.clear()
        for folder_element in wurzel.find('IgnoredFolders'):
            if folder_element.tag == 'Folder':
                self.Ignored_folders.append(folder_element.text)
                
        # find IgnoredParameters
        self.ignored_parameters.clear()
        for parameter_element in wurzel.find('IgnoredParameters'):
            if parameter_element.tag == 'Parameter':
                self.ignored_parameters.append(parameter_element.text)
        
        
    def get_filters(self):
        # get the filters
        
        # get path to xml
        original_working_dir = os.getcwd()
        new_path = self.get_new_working_path(original_working_dir)
        xml_file_path = os.path.join(new_path, 'filters.xml')

        # parse the XML data
        tree = ET.parse(xml_file_path)
        wurzel = tree.getroot()

        # delete old filters
        self.filter.clear()
        
        for element in wurzel.find('Always'):
            if element.tag == 'Filter':
                self.filter.append(element.text)
        
        if self.filter_index == 1:
            for element in wurzel.find('Induction'):
                if element.tag == 'Filter':
                    self.filter.append(element.text)
                    
        elif self.filter_index == 2:   
            for element in wurzel.find('Dynamicbuffer'):
                if element.tag == 'Filter':
                    self.filter.append(element.text)
                    
        elif self.filter_index == 3:   
            for element in wurzel.find('Orderbuffer'):
                if element.tag == 'Filter':
                    self.filter.append(element.text)  
                      
        elif self.filter_index == 4:   
            for element in wurzel.find('Matrix_Presorter'):
                if element.tag == 'Filter':
                    self.filter.append(element.text)    
                    
        elif self.filter_index == 5:   
            for element in wurzel.find('Packing'):
                if element.tag == 'Filter':
                    self.filter.append(element.text)
                    
        elif self.filter_index == 6:   
            for element in wurzel.find('Crossover'):
                if element.tag == 'Filter':
                    self.filter.append(element.text)  
        
                
    def write_first_project_to_mainfile(self):
        # write the first project to column B
        
        row_param = []
        row_value = []
        
        # 0 because its always the first project
        for parameter in self.project_list[0].standard_template.parameters:
            row_param.append(parameter.name)
            row_value.append(parameter.value)
        
        df = panda.DataFrame({'Parameter': row_param, self.project_list[0].name: row_value})
        df.to_excel(self.path_file_comparison,index=False)  
    
        
    def add_next_project_to_mainfile(self,index_in_project_list):
        # add the next project in the project_list to the next free column
        
        workbook = openpyxl.load_workbook(self.path_file_comparison)
        sheet = workbook.active
        sheet.cell(row=1, column=self.project_list[index_in_project_list].column_in_excel, value=self.project_list[index_in_project_list].name)
    
        # read every parameter from project and look where it is in the first project that is already in the excel
        for parameter in self.project_list[index_in_project_list].standard_template.parameters:
            look_up_parameter = parameter.name
            counter = 0
            look_up_parameter_found = False
            for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row, values_only=True):
                counter = counter + 1
                if row[0] == look_up_parameter:
                    sheet.cell(row=counter, column=self.project_list[index_in_project_list].column_in_excel, value=parameter.value)
                    look_up_parameter_found = True
                    break
            
            # if parameter wasnt found add it at the end
            if not look_up_parameter_found:
                next_free_row = sheet.max_row+1
                sheet.cell(row=next_free_row, column=1, value=parameter.name)
                sheet.cell(row=next_free_row, column=self.project_list[index_in_project_list].column_in_excel, value=parameter.value)
    
        workbook.save(self.path_file_comparison)
        
        
    def compare_parameters_in_file(self,file_path,compare_column_in_excel, column_in_excel):
        
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        counter = 2
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            
            value_row_master = str(row[compare_column_in_excel-1]) 
            value_row_x = str(row[column_in_excel-1])  

            if value_row_master != value_row_x:
                    cell = sheet.cell(row=counter,column = column_in_excel)
                    cell.fill = PatternFill(start_color="FF8888", end_color="FF0000", fill_type="solid")
            counter = counter + 1
        
        workbook.save(file_path)
        
        
    def write_local_changes_to_excel(self,index_in_project_list):
        
        # create poth for new excel file
        path_temp = os.path.join(self.path_target_folder,f"{self.project_list[index_in_project_list].name}.xlsx")
        
        row_param = []
        row_value = []
            
        for parameter in self.project_list[index_in_project_list].local_parameter_changes.parameters:
            row_param.append(parameter.name)
            row_value.append(parameter.value)
            
        df = panda.DataFrame({'Parameter': row_param, self.project_list[index_in_project_list].name: row_value})
        df.to_excel(path_temp,index=False) 
        
        self.format_sheet(path_temp)
        

    def copy_comparison_with_filter(self):
        src_wb = openpyxl.load_workbook(self.path_file_comparison)
        dest_wb = openpyxl.Workbook()

        src_sheet = src_wb.active
        dest_sheet = dest_wb.active
        counter = 0
        
        for row_strings in src_sheet.iter_rows(min_row=0, values_only=True):
            counter = counter + 1 
            #copy if row is in filter or its the first row / header
            if row_strings[1] in self.filter or counter == 1:
                #add row to destination sheet
                dest_sheet.append(row_strings)
                
                # copy the format and cell filling
                dest_row_index = dest_sheet.max_row
                dest_row = dest_sheet[dest_row_index]
                src_row = src_sheet[counter]
                
                for col in range(0,len(src_row)):
                    dest_row[col].font = copy(src_row [col].font)
                    dest_row[col].border = copy(src_row [col].border)
                    dest_row[col].fill = copy(src_row [col].fill)
                    dest_row[col].number_format = copy(src_row [col].number_format)
                    dest_row[col].alignment = copy(src_row [col].alignment)
            
        dest_wb.save(self.path_file_filtered_comparion)
    
    
    def cell_is_red(self,cell):
        return (
            cell.fill.bgColor.rgb == "00FF0000"
        )   
    
    
    def create_list_with_differences(self,number_of_projects):
        src_wb = openpyxl.load_workbook(self.path_file_comparison)
        src_sheet = src_wb.active
        
        dest_wb = openpyxl.Workbook()
        dest_sheet = dest_wb.active
        
        # copy header row
        for row in src_sheet.iter_rows(min_row=1, max_row=1):
            dest_sheet.append([cell.value for cell in row])
        
        # copy row if it has red filled cell
        for row in src_sheet.iter_rows(min_row=2, max_col = number_of_projects + 3, max_row=src_sheet.max_row):
            row_data = [cell.value for cell in row]
            if any(self.cell_is_red(cell) for cell in row):
                dest_sheet.append(row_data)
                
        dest_wb.save(self.path_file_differences)
    
    
    def format_parameter_column(self, file_path):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        sheet.insert_cols(2, amount=2)
        
        sheet.cell(row=1, column=1).value = 'GVL_StandardTemplate'
        sheet.cell(row=1, column=2).value = 'Device'
        sheet.cell(row=1, column=3).value = 'Parameter'
        
        counter = 2
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            if row[0] and "." in row[0]:
                parts = row[0].split(".")
                sheet.cell(row=counter, column=1).value = parts[0]
                sheet.cell(row=counter, column=2).value = parts[1]
                sheet.cell(row=counter, column=3).value = '.'.join(parts[2:])
            counter = counter + 1
            
        workbook.save(file_path)
    
    
    def format_sheet(self,file_path):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        for column_letter in "ABCDEFGH":
            sheet.column_dimensions[column_letter].width = 50
        
        font = Font(bold=True)
        alignment = Alignment(horizontal='center')
        border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

        header = sheet[1]
        # Setze die Formatierung für jede Zelle in der oberen Zeile
        for cell in header:
            cell.font = font
            cell.alignment = alignment
            cell.border = border

        workbook.save(file_path)
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
    ### currently not working
    def set_filters(self):
        
        # NOT WORKING SO FAR
        
        #load workbook
        workbook = openpyxl.load_workbook(self.path_file_comparison)
        sheet = workbook.active
       
        # set filters
        #column_B = 'B'  # Ändere dies entsprechend deiner Spalte
        #sheet.auto_filter.ref = f'{column_B}:{column_B}'

        # Setze den Filter für die gesamte Spalte
        #filter_column = FilterColumn(column_id=column_B, customFilters=[CustomFilter(operator="contains", val=val) for val in self.filter])
        #sheet.add_filter_column(filter_column)
        

        workbook.save(self.path_file_comparison)